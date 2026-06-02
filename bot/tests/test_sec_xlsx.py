"""Security: formula-injection (CWE-1236) в XLSX-выгрузке /stats.

Текст жителя (summary, address, topic, first_name, phone) и ответ
оператора попадают в ячейки XLSX через services/stats._render_workbook.
openpyxl пишет inline-строкой почти всё, КРОМЕ строки с ведущим '=' —
её он превращает в живой тег <f>, и `=HYPERLINK(...)` / `=cmd|...`
исполнится при открытии отчёта оператором в Excel. Триггер атаки:
житель кладёт формулу в суть обращения → cron-автоотчёт 1-го числа или
ручной /stats рендерит её как формулу.

Фикс: services/stats._sanitize_cell префиксует апострофом любую
user-controlled строку, начинающуюся с = + - @ \\t \\r. Эти тесты
рендерят РЕАЛЬНЫЙ workbook (не мок) и проверяют распакованный
xl/worksheets/sheet1.xml — доказательство, что тега <f> там нет.

Стенд-объекты — SimpleNamespace вместо ORM-моделей: _render_workbook
синхронный и только читает атрибуты, БД не нужна (нет гонки с
параллельными агентами по sqlite/__pycache__).
"""
from __future__ import annotations

import io
import zipfile
from datetime import datetime, timezone
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from aemr_bot.services import stats


def _user(**over):
    base = dict(
        first_name="Иван",
        phone="+79990001122",
        max_user_id=12345,
        consent_pdn_at=None,
        consent_revoked_at=None,
        subscribed_broadcast=False,
        is_blocked=False,
    )
    base.update(over)
    return SimpleNamespace(**base)


def _msg(text: str, direction: str = "from_operator"):
    return SimpleNamespace(text=text, direction=direction)


def _appeal(**over):
    base = dict(
        id=1,
        created_at=datetime(2026, 6, 1, 9, 0, tzinfo=timezone.utc),
        answered_at=None,
        status="new",
        locality="Посёлок",
        address="ул. Тестовая, 1",
        topic="Дороги",
        summary="Яма на дороге",
        messages=[],
        user=_user(),
    )
    base.update(over)
    return SimpleNamespace(**base)


def _sheet_xml(content: bytes) -> str:
    """Распаковать первый worksheet из XLSX-байтов."""
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        return z.read("xl/worksheets/sheet1.xml").decode("utf-8")


# --- единичный хелпер ---------------------------------------------------------


class TestSanitizeCell:
    @pytest.mark.parametrize(
        "raw",
        [
            '=HYPERLINK("http://evil","click")',
            "=1+1",
            "+1+1",
            "-1+1",
            "@SUM(1)",
            "\tcmd",
            "\rcmd",
            "=cmd|'/c calc'!A1",
        ],
    )
    def test_prefixes_apostrophe_for_formula_triggers(self, raw: str) -> None:
        out = stats._sanitize_cell(raw)
        assert out == "'" + raw
        assert out.startswith("'")

    @pytest.mark.parametrize(
        "raw",
        ["Иван", "Яма на дороге", "ул. Ленина, 5", "", "100", "  =later"],
    )
    def test_passes_through_safe_strings(self, raw: str) -> None:
        # Ведущий пробел не делает строку формулой — Excel не вычисляет.
        assert stats._sanitize_cell(raw) == raw

    def test_non_strings_unchanged(self) -> None:
        assert stats._sanitize_cell(123) == 123
        assert stats._sanitize_cell(None) is None
        assert stats._sanitize_cell(3.14) == 3.14


# --- рендер реального workbook ------------------------------------------------


class TestNoLiveFormulaInWorkbook:
    def test_hyperlink_in_summary_is_not_a_formula(self) -> None:
        """Канонический PoC из находки: =HYPERLINK(...) в сути → нет <f>."""
        payload = '=HYPERLINK("http://attacker.example/x","click")'
        content = stats._render_workbook([_appeal(summary=payload)])
        xml = _sheet_xml(content)
        assert "<f>" not in xml
        assert "HYPERLINK" in xml  # значение сохранено, просто как текст

    def test_all_user_fields_neutralised(self) -> None:
        """Каждое user-controlled поле с формулой → ни одного тега <f>."""
        appeal = _appeal(
            summary="=HYPERLINK(\"http://e/s\",\"s\")",
            address="=cmd|'/c calc'!A1",
            topic="+1+1",
            locality="-2+2",
            user=_user(
                first_name="=1+2",
                phone="@SUM(A1:A9)",
            ),
            messages=[_msg("=HYPERLINK(\"http://e/r\",\"r\")")],
        )
        content = stats._render_workbook([appeal])
        xml = _sheet_xml(content)
        assert "<f>" not in xml

    def test_legitimate_text_workbook_has_no_formula(self) -> None:
        """Сохранение поведения: обычное обращение не ломается."""
        content = stats._render_workbook([_appeal()])
        xml = _sheet_xml(content)
        assert "<f>" not in xml

    def test_roundtrip_value_carries_apostrophe_guard(self) -> None:
        """После сохранения ячейка читается как текст (с ведущим '),
        НЕ как вычисленная формула. Это и есть нейтрализация."""
        payload = '=HYPERLINK("http://attacker.example/x","click")'
        content = stats._render_workbook([_appeal(summary=payload)])
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # Заголовок в строке 1, данные с строки 2. "Суть" — колонка 9.
        cell = ws.cell(row=2, column=9)
        assert cell.data_type != "f", "ячейка не должна быть формулой"
        assert cell.value == "'" + payload

    def test_operator_reply_neutralised_roundtrip(self) -> None:
        """Ответ оператора (колонка 11) — тоже user-controlled."""
        payload = "=1+1"
        appeal = _appeal(messages=[_msg(payload, direction="from_operator")])
        content = stats._render_workbook([appeal])
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        cell = ws.cell(row=2, column=11)
        assert cell.data_type != "f"
        assert cell.value == "'" + payload
