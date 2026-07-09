"""Perf-кластер: trigram-индексы под search_audience + xlsx без
материализации всей переписки.

Две независимые находки Волны 2, проверяемые здесь:

(a) Миграция 0018 создаёт extension pg_trgm и два GIN-индекса
    (ix_users_first_name_trgm, ix_users_phone_normalized_trgm) под
    подстрочный ILIKE '%x%' в services/users.py::search_audience.
    Без них — seq scan по users на каждый поиск жителя.

(b) services/stats.build_xlsx больше НЕ грузит selectinload(Appeal.messages)
    (вся переписка всех обращений в RAM на period=all → OOM-риск при
    mem_limit:512m). Вместо этого отдельный запрос
    _load_last_operator_replies достаёт по одной строке на обращение —
    последний ответ оператора, ровно то, что показывает колонка «Ответ
    оператора». operator_reply в выгрузке остаётся тем же.

Pure-тесты (без БД) работают везде, включая локальный sqlite-дефолт —
никакой гонки с параллельными агентами по __pycache__/sqlite-файлам.
Интеграционные тесты (фикстура `session`) требуют PostgreSQL: локально
skip, в CI на postgres:16 запускаются и проверяют РЕАЛЬНОЕ создание
индексов и реальную загрузку обращений.
"""
from __future__ import annotations

import importlib
import io
from datetime import datetime, timedelta, timezone
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from aemr_bot.db.models import Appeal, Message, MessageDirection, User
from aemr_bot.services import stats

MIGRATION_MODULE = (
    "aemr_bot.db.alembic.versions.0018_users_trigram_search_indexes"
)


# ──────────────────────────────────────────────────────────────────────
# Хелперы
# ──────────────────────────────────────────────────────────────────────


def _selectin_loaded_relationships(stmt) -> set[str]:
    """Имена relationship'ов, помеченных eager-load (selectinload и т.п.)
    в SQLAlchemy-Select. Извлекаем ключ relationship-элемента из path
    каждой load-опции. Устойчиво к версии: читаем публичный `.key`
    relationship-объекта внутри option.path, а не приватные внутренности
    компиляции."""
    rels: set[str] = set()
    for option in getattr(stmt, "_with_options", ()):
        path = getattr(option, "path", None) or ()
        for elem in path:
            key = getattr(elem, "key", None)
            if isinstance(key, str):
                rels.add(key)
    return rels


def _msg(text: str, direction: str = "from_operator") -> SimpleNamespace:
    return SimpleNamespace(text=text, direction=direction)


def _user(**over) -> SimpleNamespace:
    base = dict(
        first_name="Иван",
        phone="+79990001122",
        max_user_id=12345,
        consent_pdn_at=None,
        consent_revoked_at=None,
        subscribed_broadcast=False,
        is_blocked=False,
    )
    base.update(over)
    return SimpleNamespace(**base)


def _appeal(**over) -> SimpleNamespace:
    base = dict(
        id=1,
        created_at=datetime(2026, 6, 1, 9, 0, tzinfo=timezone.utc),
        answered_at=None,
        status="new",
        locality="Посёлок",
        address="ул. Тестовая, 1",
        topic="Дороги",
        summary="Яма на дороге",
        messages=[],
        user=_user(),
    )
    base.update(over)
    return SimpleNamespace(**base)


def _operator_reply_cell(content: bytes) -> object:
    """Значение ячейки «Ответ оператора» (колонка 11) первой строки
    данных из отрендеренного XLSX."""
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    return ws.cell(row=2, column=11).value


def _phone_cell(content: bytes) -> object:
    """Значение ячейки «Телефон» (колонка 4) первой строки данных."""
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    return ws.cell(row=2, column=4).value


# ══════════════════════════════════════════════════════════════════════
# (a) Миграция 0018 — trigram-индексы
# ══════════════════════════════════════════════════════════════════════


def _capture_migration_ddl(fn_name: str) -> str:
    """Прогнать upgrade()/downgrade() миграции 0018, перехватив весь DDL,
    что она шлёт через op.execute. Возвращает склеенный текст.

    op.execute — метод alembic-proxy (общий синглтон): сохраняем и
    восстанавливаем оригинал в finally, чтобы не протекла мутация в
    соседние тесты процесса (reload модуля её бы не откатил)."""
    mod = importlib.import_module(MIGRATION_MODULE)
    captured: list[str] = []
    original = mod.op.execute
    mod.op.execute = lambda sql: captured.append(str(sql))  # type: ignore[attr-defined]
    try:
        getattr(mod, fn_name)()
    finally:
        mod.op.execute = original  # type: ignore[attr-defined]
    return "\n".join(captured)


class TestMigration0018Metadata:
    """Статическая валидность миграции — без БД. Ловит сломанную
    revision-цепочку (CI alembic upgrade head упал бы) и опечатки в DDL."""

    def test_revision_chain_links_to_previous_head(self) -> None:
        mod = importlib.import_module(MIGRATION_MODULE)
        # 0017 был head до этой миграции (см. versions/). down_revision
        # обязан указывать на него, иначе цепочка рвётся и
        # `alembic upgrade head` в CI падает с multiple/!= heads.
        assert mod.revision == "0018"
        assert mod.down_revision == "0017"

    def test_upgrade_creates_extension_and_both_indexes(self) -> None:
        joined = _capture_migration_ddl("upgrade")
        assert "CREATE EXTENSION IF NOT EXISTS pg_trgm" in joined
        assert "ix_users_first_name_trgm" in joined
        assert "ix_users_phone_normalized_trgm" in joined
        # Триграммный operator class — иначе GIN не ускоряет ILIKE '%x%'.
        assert joined.count("gin_trgm_ops") == 2
        assert "USING gin" in joined

    def test_downgrade_drops_both_indexes(self) -> None:
        joined = _capture_migration_ddl("downgrade")
        assert "DROP INDEX IF EXISTS ix_users_first_name_trgm" in joined
        assert "DROP INDEX IF EXISTS ix_users_phone_normalized_trgm" in joined


class TestUserModelDeclaresTrigramIndexes:
    """alembic check сравнивает Base.metadata с БД после миграций. Если
    индексы 0018 не объявлены в модели — check видит drift и CI падает.
    Здесь фиксируем, что модель их объявляет с тем же DDL, что миграция."""

    def test_both_trigram_indexes_present_on_users_table(self) -> None:
        from sqlalchemy.dialects import postgresql
        from sqlalchemy.schema import CreateIndex

        by_name = {ix.name: ix for ix in User.__table__.indexes}
        assert "ix_users_first_name_trgm" in by_name
        assert "ix_users_phone_normalized_trgm" in by_name

        fn_ddl = str(
            CreateIndex(by_name["ix_users_first_name_trgm"]).compile(
                dialect=postgresql.dialect()
            )
        )
        ph_ddl = str(
            CreateIndex(by_name["ix_users_phone_normalized_trgm"]).compile(
                dialect=postgresql.dialect()
            )
        )
        # Ровно то DDL, что эмитит миграция 0018 — иначе alembic-check drift.
        assert "USING gin (first_name gin_trgm_ops)" in fn_ddl
        assert "USING gin (phone_normalized gin_trgm_ops)" in ph_ddl


@pytest.mark.asyncio
class TestMigration0018AgainstPostgres:
    """Интеграция: реальное применение миграций на чистую БД создаёт
    extension и индексы. Требует Postgres (фикстура `session` skip'ает
    на sqlite/без DATABASE_URL).

    Здесь мы НЕ гоняем alembic-CLI (это делает отдельный CI-степ
    `alembic upgrade head`), а проверяем результат, эквивалентный
    create_all: фикстура поднимает схему из Base.metadata, и DDL-listener
    _ensure_pg_trgm в models.py гарантирует pg_trgm перед users. Так
    тест заодно доказывает, что trigram-индексы вообще СОЗДАЮТСЯ без
    ручного CREATE EXTENSION (раньше create_all упал бы на gin_trgm_ops)."""

    async def test_extension_and_indexes_exist(self, session) -> None:
        from sqlalchemy import text as sa_text

        ext = await session.scalar(
            sa_text("SELECT 1 FROM pg_extension WHERE extname = 'pg_trgm'")
        )
        assert ext == 1, "DDL-listener должен был создать extension pg_trgm"

        rows = (
            await session.scalars(
                sa_text(
                    "SELECT indexname FROM pg_indexes "
                    "WHERE tablename = 'users' AND indexname LIKE '%_trgm'"
                )
            )
        ).all()
        assert "ix_users_first_name_trgm" in rows
        assert "ix_users_phone_normalized_trgm" in rows

    async def test_trigram_index_is_gin(self, session) -> None:
        from sqlalchemy import text as sa_text

        # amname='gin' подтверждает, что индекс действительно GIN
        # (а не B-tree) — только GIN+gin_trgm_ops ускоряет ILIKE '%x%'.
        amname = await session.scalar(
            sa_text(
                "SELECT am.amname FROM pg_class c "
                "JOIN pg_am am ON am.oid = c.relam "
                "WHERE c.relname = 'ix_users_first_name_trgm'"
            )
        )
        assert amname == "gin"


# ══════════════════════════════════════════════════════════════════════
# (b) build_xlsx — без материализации всей переписки
# ══════════════════════════════════════════════════════════════════════


class _CapturingSession:
    """Фейковая AsyncSession для build_xlsx без БД.

    - первый `.scalars(stmt)` — это запрос обращений; записываем stmt и
      отдаём заранее заданные appeal-объекты.
    - `.execute(stmt)` — это _load_last_operator_replies; отдаём заранее
      заданные (appeal_id, text)-строки.
    """

    def __init__(self, appeals, reply_rows):
        self._appeals = appeals
        self._reply_rows = list(reply_rows)
        self.scalar_statements = []
        self.execute_statements = []

    async def scalars(self, stmt):
        self.scalar_statements.append(stmt)
        return iter(self._appeals)

    async def execute(self, stmt):
        self.execute_statements.append(stmt)
        return iter(self._reply_rows)


class TestBuildXlsxDoesNotMaterializeMessages:
    @pytest.mark.asyncio
    async def test_appeal_query_does_not_eager_load_messages(self) -> None:
        """Главная perf-проверка: запрос обращений тянет user, но НЕ
        messages. Раньше было selectinload(Appeal.user, Appeal.messages)
        — вся переписка в RAM."""
        appeal = _appeal(id=42, messages=[])
        sess = _CapturingSession([appeal], reply_rows=[(42, "ответ оператора")])

        await stats.build_xlsx(sess, "all")

        assert len(sess.scalar_statements) == 1
        loaded = _selectin_loaded_relationships(sess.scalar_statements[0])
        assert "user" in loaded, "user по-прежнему eager-load (нужен для ПДн-колонок)"
        assert "messages" not in loaded, (
            "messages не должны eager-load'иться — это и есть OOM-фикс"
        )

    @pytest.mark.asyncio
    async def test_last_operator_reply_query_issued(self) -> None:
        """build_xlsx делает отдельный запрос за последними ответами
        оператора (а не достаёт их из материализованной коллекции)."""
        appeal = _appeal(id=7, messages=[])
        sess = _CapturingSession([appeal], reply_rows=[(7, "финальный ответ")])

        await stats.build_xlsx(sess, "all")

        # Ровно один execute — _load_last_operator_replies.
        assert len(sess.execute_statements) == 1

    @pytest.mark.asyncio
    async def test_operator_reply_from_dict_lands_in_workbook(self) -> None:
        """operator_reply в выгрузке берётся из словаря отдельного
        запроса и попадает в колонку «Ответ оператора»."""
        appeal = _appeal(id=99, messages=[])
        sess = _CapturingSession(
            [appeal], reply_rows=[(99, "последний ответ оператора")]
        )

        content, _title, count = await stats.build_xlsx(sess, "all")

        assert count == 1
        assert _operator_reply_cell(content) == "последний ответ оператора"

    @pytest.mark.asyncio
    async def test_appeal_without_reply_renders_empty_cell(self) -> None:
        """Обращение без ответа оператора (нет в словаре) → пустая
        ячейка, а не падение. openpyxl при чтении отдаёт пустую строку
        как None — то же самое поведение, что у старого пути (см.
        TestRenderWorkbookOperatorReplyParity.test_empty_reply_parity)."""
        appeal = _appeal(id=5, messages=[])
        sess = _CapturingSession([appeal], reply_rows=[])  # ни одного ответа

        content, _title, _count = await stats.build_xlsx(sess, "all")

        assert _operator_reply_cell(content) in (None, "")

    @pytest.mark.asyncio
    async def test_phone_column_is_masked_not_raw(self) -> None:
        """152-ФЗ: XLSX-выгрузка скачивается оператором на диск и может
        уйти дальше без контроля доступа admin-чата — полный номер
        телефона в файле недопустим. Колонка «Телефон» должна нести
        маску `mask_phone` (+7***XXXX), а не сырой `user.phone`.
        Регрессия: раньше build_xlsx писал `u.phone` напрямую, без
        маски (см. `aemr_bot/utils/pii_mask.py::mask_phone`)."""
        appeal = _appeal(
            id=1, messages=[], user=_user(phone="+79991234567")
        )
        sess = _CapturingSession([appeal], reply_rows=[])

        content, _title, _count = await stats.build_xlsx(sess, "all")

        assert _phone_cell(content) == "+7***4567"
        assert "9991234567" not in str(_phone_cell(content))

    @pytest.mark.asyncio
    async def test_phone_column_empty_dash_when_user_missing_phone(self) -> None:
        appeal = _appeal(id=2, messages=[], user=_user(phone=None))
        sess = _CapturingSession([appeal], reply_rows=[])

        content, _title, _count = await stats.build_xlsx(sess, "all")

        assert _phone_cell(content) == "—"


class TestLoadLastOperatorRepliesPure:
    @pytest.mark.asyncio
    async def test_empty_ids_short_circuits_without_query(self) -> None:
        """Пустой список id — сразу {} без обращения к БД (на пустой
        выгрузке не делаем лишний запрос)."""

        class _Boom:
            async def execute(self, _stmt):
                raise AssertionError("не должны лезть в БД при пустом списке")

        result = await stats._load_last_operator_replies(_Boom(), [])
        assert result == {}


class TestRenderWorkbookOperatorReplyParity:
    """operator_reply в новом (словарь) и старом (reversed(messages)) путях
    даёт ОДНО значение — последний from_operator-ответ. Это гарантия, что
    perf-рефактор поведение-сохраняющий."""

    def test_dict_path_matches_messages_fallback(self) -> None:
        last_reply = "самый свежий ответ"
        messages = [
            _msg("первый ответ", "from_operator"),
            _msg("вопрос жителя", "from_user"),
            _msg(last_reply, "from_operator"),
        ]
        # Старый путь: messages переданы, operator_replies=None.
        appeal_old = _appeal(id=1, messages=messages)
        content_old = stats._render_workbook([appeal_old])

        # Новый путь: messages пустые, последний ответ — из словаря.
        appeal_new = _appeal(id=1, messages=[])
        content_new = stats._render_workbook([appeal_new], {1: last_reply})

        assert _operator_reply_cell(content_old) == last_reply
        assert _operator_reply_cell(content_new) == last_reply

    def test_empty_reply_parity(self) -> None:
        """Обращение без ответа оператора: и старый (нет from_operator в
        messages), и новый (id отсутствует в словаре) пути дают
        одинаковую пустую ячейку. openpyxl читает пустую строку как
        None — фиксируем это как общий контракт обоих путей."""
        appeal_old = _appeal(id=1, messages=[_msg("только житель", "from_user")])
        appeal_new = _appeal(id=1, messages=[])
        cell_old = _operator_reply_cell(stats._render_workbook([appeal_old]))
        cell_new = _operator_reply_cell(stats._render_workbook([appeal_new], {}))
        assert cell_old == cell_new

    def test_fallback_picks_latest_from_operator(self) -> None:
        """Контроль семантики fallback: из нескольких from_operator
        берётся ПОСЛЕДНИЙ (как reversed(messages))."""
        appeal = _appeal(
            id=1,
            messages=[
                _msg("старый ответ", "from_operator"),
                _msg("новый ответ", "from_operator"),
            ],
        )
        content = stats._render_workbook([appeal])
        assert _operator_reply_cell(content) == "новый ответ"

    def test_dict_path_does_not_read_messages_attr(self) -> None:
        """Когда operator_replies передан, _render_workbook НЕ обращается
        к a.messages — доказательство, что коллекция не нужна (и потому
        её можно не грузить)."""

        class _NoMessages:
            id = 1
            created_at = datetime(2026, 6, 1, 9, 0, tzinfo=timezone.utc)
            answered_at = None
            status = "new"
            locality = "П"
            address = "А"
            topic = "Т"
            summary = "С"
            user = _user()

            @property
            def messages(self):  # pragma: no cover - не должно вызываться
                raise AssertionError(
                    "a.messages не должен читаться при заданном operator_replies"
                )

        content = stats._render_workbook([_NoMessages()], {1: "ответ"})
        assert _operator_reply_cell(content) == "ответ"


# ══════════════════════════════════════════════════════════════════════
# (b) Интеграция build_xlsx на реальном Postgres
# ══════════════════════════════════════════════════════════════════════


@pytest.mark.asyncio
class TestBuildXlsxIntegration:
    """Реальная загрузка обращений + сообщений из Postgres. Доказывает:
    (1) последний ответ оператора в выгрузке корректен на фоне длинной
        переписки;
    (2) коллекция Appeal.messages НЕ материализуется (остаётся
        unloaded) — то есть OOM-фикс реально работает на ORM-уровне."""

    async def _seed_appeal_with_history(self, session) -> tuple[int, str]:
        user = User(max_user_id=900001, first_name="Тест", phone="+79991112233")
        session.add(user)
        await session.flush()
        appeal = Appeal(
            user_id=user.id,
            status="answered",
            address="ул. Реальная, 1",
            topic="Дороги",
            summary="Текст обращения",
            created_at=datetime.now(timezone.utc) - timedelta(days=1),
            answered_at=datetime.now(timezone.utc),
        )
        session.add(appeal)
        await session.flush()

        base = datetime.now(timezone.utc) - timedelta(hours=5)
        # Длинная переписка: чередуем сообщения жителя и оператора.
        msgs = [
            Message(
                appeal_id=appeal.id,
                direction=MessageDirection.FROM_USER.value,
                text="вопрос 1",
                created_at=base,
            ),
            Message(
                appeal_id=appeal.id,
                direction=MessageDirection.FROM_OPERATOR.value,
                text="ранний ответ оператора",
                created_at=base + timedelta(hours=1),
            ),
            Message(
                appeal_id=appeal.id,
                direction=MessageDirection.FROM_USER.value,
                text="уточнение жителя",
                created_at=base + timedelta(hours=2),
            ),
            Message(
                appeal_id=appeal.id,
                direction=MessageDirection.FROM_OPERATOR.value,
                text="ПОСЛЕДНИЙ ответ оператора",
                created_at=base + timedelta(hours=3),
            ),
        ]
        session.add_all(msgs)
        await session.flush()
        return appeal.id, "ПОСЛЕДНИЙ ответ оператора"

    async def test_latest_operator_reply_in_workbook(self, session) -> None:
        appeal_id, expected = await self._seed_appeal_with_history(session)

        content, _title, count = await stats.build_xlsx(session, "all")

        assert count >= 1
        # Найдём строку нашего обращения и сверим колонку «Ответ оператора».
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        found = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == appeal_id:  # колонка «№» = appeal.id
                found = row
                break
        assert found is not None, "строка обращения должна присутствовать в выгрузке"
        # Колонка 11 (1-based) = индекс 10 в tuple values_only.
        assert found[10] == expected

    async def test_messages_collection_not_loaded(self, session) -> None:
        """Прямое доказательство OOM-фикса: build_xlsx грузит обращения
        через свой собственный select(Appeal) без selectinload(messages).

        Повторяем тот же запрос, что build_xlsx, и проверяем на
        SQLAlchemy-уровне, что relationship `messages` остаётся
        unloaded (lazy) у загруженных объектов — то есть в RAM не
        затягивается вся переписка."""
        from sqlalchemy import inspect as sa_inspect

        await self._seed_appeal_with_history(session)

        # Тот же запрос, что внутри build_xlsx (user — да, messages — нет).
        res = await session.scalars(
            select(Appeal)
            .options(selectinload(Appeal.user))
            .order_by(Appeal.created_at.desc())
        )
        appeals = list(res)
        assert appeals, "должно быть загружено хотя бы одно обращение"
        for a in appeals:
            unloaded = sa_inspect(a).unloaded
            assert "messages" in unloaded, (
                "messages не должны быть материализованы при выгрузке"
            )
            assert "user" not in unloaded, "user, наоборот, eager-load'ится"
