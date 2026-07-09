import asyncio
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from zoneinfo import ZoneInfo

from aemr_bot.config import settings
from aemr_bot.db.models import Appeal, AppealStatus, Message, MessageDirection
from aemr_bot.services import sla as sla_service
from aemr_bot.utils.pii_mask import mask_phone

TZ = ZoneInfo(settings.timezone)


VALID_PERIODS = ("today", "week", "month", "quarter", "half_year", "year", "all")


# Защита от CSV/formula-injection в XLSX. openpyxl сам пишет inline-строкой
# (не формулой) почти всё, но строку с ведущим '=' превращает в живой тег
# <f> — то есть `=HYPERLINK(...)` / `=cmd|...` из текста жителя (summary,
# address, topic, first_name) или ответа оператора исполнится при открытии
# в Excel. Ведущие + - @ \t \r остальные ридеры (LibreOffice, Excel при
# ре-сейве) тоже трактуют как начало формулы. Fail-closed: для любого
# user-controlled значения, начинающегося с этих символов, префиксуем
# апострофом — Excel показывает текст как есть и не вычисляет.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value):
    """Нейтрализует formula-injection для user-controlled ячеек XLSX.

    Строку, начинающуюся с потенциального формула-триггера, префиксует
    апострофом. Не-строки (int id, числа, None) возвращает без изменений —
    они не могут нести формулу. Поведение для легитимного текста не
    меняется: '=' в начале у обычного обращения встречается крайне редко,
    а апостроф невидим в ячейке (Excel-конвенция «текст as-is»).
    """
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGERS):
        return "'" + value
    return value


# Потолок строк в одной XLSX-выгрузке. На `period="all"` без лимита
# годовой архив (10k+ обращений × N сообщений каждое, с selectinload)
# тянет всё в RAM разом — при mem_limit:512m это риск OOM. 10000 строк
# — заведомо больше реальной операторской потребности, а сам XLSX с
# таким числом строк уже неюзабелен для чтения. При превышении —
# берём свежайшие и помечаем в заголовке.
_XLSX_ROW_CAP = 10000


def period_window(period: str) -> tuple[datetime | None, datetime, str]:
    """Окно периода для /stats. Возвращает (start_utc, end_utc, title).

    Для `all` start_utc=None — выгрузка идёт без нижнего фильтра по дате,
    то есть «за всё время существования бота». Все остальные значения
    дают конкретный start.
    """
    now = datetime.now(TZ)
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if period == "today":
        start = midnight
        title = f"за сегодня {start:%d.%m.%Y}"
    elif period == "week":
        start = midnight - timedelta(days=7)
        title = f"за 7 дней с {start:%d.%m.%Y}"
    elif period == "month":
        start = midnight - timedelta(days=30)
        title = f"за 30 дней с {start:%d.%m.%Y}"
    elif period == "quarter":
        start = midnight - timedelta(days=90)
        title = f"за квартал с {start:%d.%m.%Y}"
    elif period == "half_year":
        start = midnight - timedelta(days=183)
        title = f"за полгода с {start:%d.%m.%Y}"
    elif period == "year":
        start = midnight - timedelta(days=365)
        title = f"за год с {start:%d.%m.%Y}"
    elif period == "all":
        return None, now.astimezone(timezone.utc), "за всё время"
    else:
        raise ValueError(f"Unknown period: {period}")
    return start.astimezone(timezone.utc), now.astimezone(timezone.utc), title


async def build_xlsx(session: AsyncSession, period: str) -> tuple[bytes, str, int]:
    start, end, title = period_window(period)
    # Берём свежайшие в пределах окна, с потолком _XLSX_ROW_CAP.
    # `+1` — чтобы детектировать факт обрезки, не делая отдельный count.
    #
    # ВАЖНО (perf): НЕ грузим selectinload(Appeal.messages). На period=all
    # это тянуло в RAM всю переписку всех обращений (годовой архив: 10k
    # обращений × N сообщений), хотя выгрузке нужен лишь последний ответ
    # оператора по каждому обращению (колонка «Ответ оператора»). При
    # mem_limit:512m полная материализация messages — риск OOM. Вместо
    # этого отдельным запросом достаём словарь appeal_id → последний
    # operator-reply (по одной строке на обращение, см.
    # _load_last_operator_replies).
    query = (
        select(Appeal)
        .options(selectinload(Appeal.user))
        .where(Appeal.created_at <= end)
        .order_by(Appeal.created_at.desc())
        .limit(_XLSX_ROW_CAP + 1)
    )
    if start is not None:
        query = query.where(Appeal.created_at >= start)
    res = await session.scalars(query)
    appeals = list(res)
    truncated = len(appeals) > _XLSX_ROW_CAP
    if truncated:
        appeals = appeals[:_XLSX_ROW_CAP]
        title = f"{title} (показаны последние {_XLSX_ROW_CAP})"
    # Возвращаем хронологический порядок (запрос был DESC ради лимита).
    appeals.reverse()

    # Только последний ответ оператора на каждое обращение — ровно то,
    # что читает колонка «Ответ оператора». Одна строка messages на
    # обращение вместо всей переписки.
    operator_replies = await _load_last_operator_replies(
        session, [a.id for a in appeals]
    )

    # Построение workbook — синхронный CPU/IO-bound код (openpyxl).
    # На потолке в 10k строк это ощутимо; выносим в поток, чтобы не
    # блокировать event-loop бота на время генерации отчёта.
    content = await asyncio.to_thread(_render_workbook, appeals, operator_replies)
    return content, title, len(appeals)


async def _load_last_operator_replies(
    session: AsyncSession, appeal_ids: list[int]
) -> dict[int, str | None]:
    """Словарь appeal_id → текст ПОСЛЕДНЕГО ответа оператора.

    Заменяет загрузку всей коллекции `Appeal.messages` для XLSX-выгрузки.
    Раньше `_render_workbook` брал из полной переписки лишь последний
    `from_operator`-ответ (`next(... reversed(a.messages) ...)`), а
    остальные сообщения загружались зря — на period=all это вся история
    бота в RAM.

    Здесь оконная функция row_number() поверх messages, отфильтрованных
    по direction='from_operator' и нужным appeal_id, отдаёт по одной
    строке на обращение — самую свежую. Порядок `created_at DESC, id DESC`
    повторяет семантику старого `reversed(messages)` (messages
    отсортированы по created_at ASC), а `id DESC` добавлен как
    детерминированный тай-брейк при совпадении created_at.

    Обращения без ответа оператора в словарь не попадают —
    `_render_workbook` подставит пустую строку через `.get()`.
    """
    if not appeal_ids:
        return {}
    rn = (
        func.row_number()
        .over(
            partition_by=Message.appeal_id,
            order_by=(Message.created_at.desc(), Message.id.desc()),
        )
        .label("rn")
    )
    ranked = (
        select(Message.appeal_id, Message.text, rn)
        .where(
            Message.appeal_id.in_(appeal_ids),
            Message.direction == MessageDirection.FROM_OPERATOR.value,
        )
        .subquery()
    )
    rows = await session.execute(
        select(ranked.c.appeal_id, ranked.c.text).where(ranked.c.rn == 1)
    )
    return {appeal_id: text for appeal_id, text in rows}


def _render_workbook(
    appeals: list[Appeal],
    operator_replies: dict[int, str | None] | None = None,
) -> bytes:
    """Синхронная сборка XLSX. Работает только с уже загруженными
    данными (selectinload отработал в build_xlsx) — сессию не трогает,
    поэтому безопасно вызывать из asyncio.to_thread.

    `operator_replies` — словарь appeal_id → последний ответ оператора,
    подготовленный build_xlsx через _load_last_operator_replies (вместо
    материализации всей переписки). Если не передан (None) — поведение
    откатывается к чтению `a.messages`: так продолжают работать прямые
    вызовы _render_workbook (например, security-тесты formula-injection,
    собирающие appeal со списком .messages вручную)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения"

    headers = [
        "№",
        "Создано",
        "Имя",
        "Телефон",
        "max_user_id",
        "Населённый пункт",
        "Адрес",
        "Тематика",
        "Суть",
        "Статус",
        "Ответ оператора",
        "Время ответа, ч",
        "В SLA (4ч)",
        "Согласие на ПДн",
        "Согласие отозвано",
        "Подписан на рассылку",
        "Заблокирован",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _fmt_dt(dt) -> str:
        if dt is None:
            return ""
        return dt.astimezone(TZ).strftime("%d.%m.%Y %H:%M")

    for a in appeals:
        if operator_replies is not None:
            # Оптимизированный путь build_xlsx: последний ответ оператора
            # уже выбран отдельным запросом. Обращения без ответа в
            # словаре отсутствуют → пустая строка.
            operator_reply = operator_replies.get(a.id)
        else:
            # Fallback для прямых вызовов (messages загружены целиком).
            operator_reply = next(
                (m.text for m in reversed(a.messages) if m.direction == "from_operator"),
                None,
            )
        if a.answered_at and a.created_at:
            # elapsed_hours — календарное время до ответа (факт «сколько
            # реально прошло», интересен сам по себе). "В SLA" же —
            # оценка просрочки и считается по РАБОЧЕМУ времени
            # (services/sla.py), той же логикой, что и live-напоминалки
            # cron'а: обращение, поступившее в пятницу вечером и
            # отвеченное в понедельник утром, календарно «висело» почти
            # трое суток, но по рабочим часам могло уложиться в SLA.
            elapsed = (a.answered_at - a.created_at).total_seconds()
            elapsed_hours = round(elapsed / 3600, 2)
            overdue = sla_service.is_overdue(
                a.created_at, a.answered_at, settings.sla_response_hours
            )
            in_sla = "нет" if overdue else "да"
        else:
            elapsed_hours = None
            in_sla = ""
        u = a.user
        # _sanitize_cell — на всех ячейках с текстом жителя/оператора
        # (formula-injection guard, см. _sanitize_cell). Числа и
        # программно-сформированные строки (даты, статус, флаги да/нет)
        # формулу нести не могут — их не оборачиваем.
        #
        # Телефон — через mask_phone (152-ФЗ): XLSX-выгрузка скачивается
        # оператором на диск и может уйти дальше без контроля доступа
        # admin-чата, поэтому полный номер сюда попадать не должен —
        # та же маска «+7***1234», что и в admin-уведомлениях/выборках
        # (services/admin_events.py, handlers/admin_audience.py).
        ws.append([
            a.id,
            _fmt_dt(a.created_at),
            _sanitize_cell(u.first_name if u else ""),
            mask_phone(u.phone if u else None),
            u.max_user_id if u else "",
            _sanitize_cell(a.locality or ""),
            _sanitize_cell(a.address or ""),
            _sanitize_cell(a.topic or ""),
            _sanitize_cell(a.summary or ""),
            _status_label(a.status),
            _sanitize_cell(operator_reply or ""),
            elapsed_hours if elapsed_hours is not None else "",
            in_sla,
            _fmt_dt(u.consent_pdn_at) if u else "",
            _fmt_dt(u.consent_revoked_at) if u else "",
            "да" if (u and u.subscribed_broadcast) else "нет",
            "да" if (u and u.is_blocked) else "нет",
        ])

    widths = [6, 18, 16, 18, 14, 28, 36, 24, 60, 18, 60, 14, 12, 18, 18, 14, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _status_label(status: str) -> str:
    return {
        AppealStatus.NEW.value: "Новое",
        AppealStatus.IN_PROGRESS.value: "В работе",
        AppealStatus.ANSWERED.value: "Завершено",
        AppealStatus.CLOSED.value: "Закрыто",
    }.get(status, status)
